#! /usr/bin/python3

from openpyxl import load_workbook

wb = load_workbook('data/sheet1.xlsx')
sheet = wb.active

count = 1622
c = 2

for r in range(3,count+1):
    name = sheet.cell(row=r, column=c)

    # process

    # time
    cell = sheet.cell(row=r, column=5)
    cell.value = name.value
    # phone
    cell = sheet.cell(row=r, column=6)
    cell.value = '010-139'
    # email
    cell = sheet.cell(row=r, column=7)
    cell.value = 'example@test.com'
    # address
    cell = sheet.cell(row=r, column=8)
    cell.value = '北京市'
    # industry
    cell = sheet.cell(row=r, column=9)
    cell.value = '计算机'

wb.save('data/sheet1-save.xlsx')
