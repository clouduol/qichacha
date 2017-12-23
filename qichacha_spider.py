#! /usr/bin/python3

import requests
import re
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import time
import traceback
import random

# convert cookies to dict
# copy valid cookies form browser to data/cookies first
def get_cookie():
    cookies={}
    with open('data/cookies','r') as f:
        for line in f.read().split(';'):
            name,value=line.strip().split('=',1)  #1代表只分割一次
            cookies[name]=value 
    return cookies

def get_infos(s,firm_name):
    # return dict
    infos = {"time":"", "phone":"", "email":"", "address":"", "industry":""}
    err = ""

    try:
	    search_url='http://www.qichacha.com/search?key='+firm_name
	    # search firm name
	    time.sleep(random.randint(15,20))
	    r=s.get(search_url)
	    # print(r.text)
	
	    soup=BeautifulSoup(r.text, "lxml")
	    table=soup.find('table', class_="m_srchList")
	    #if table == None:
	    #    print(firm_name+": no firm found")
	    #    exit()
	    td=table.find('tbody').find('tr').find('td').find_next_sibling()
	    partial_url=td.find('a')["href"]
	    current_p=td.find('p')
	    span_time=current_p.find('span').find_next_sibling()
	    infos["time"]=span_time.string.split("：")[1].strip("\n ")
	
	    #re_firm=r'a href="(/firm.*)" target="_blank" class="ma_h1"'
	    #partial_urls=re.findall(re_firm, r.text)
	    #if len(partial_urls) == 0:
	    #    print(firm_name+": no firm found")
	    #    return infos
	    ## use first matched firm
	    #firm_url="http://"+host+partial_urls[0]
	    ## print(firm_url)

	    #if partial_url == None:
	    #    print(firm_name+": no firm found")
	    #    exit()
	    firm_url="http://"+host+partial_url
	
	    # get firm info page
	    time.sleep(random.randint(15,20))
	    r=s.get(firm_url)
	    # print(r.text)
	    soup=BeautifulSoup(r.text,"lxml")
	    # first block
	    span_phone=soup.find('span', text="电话：")
	    if span_phone != None:
	        strtmp=span_phone.find_next_sibling().find_next().string
	        if strtmp != None:
	            strtmp = strtmp.strip("\n ")
	            infos["phone"]=strtmp
	    span_email=soup.find('span', text="邮箱：")
	    if span_email != None:
	        strtmp=span_email.find_next_sibling().find_next().string
	        if strtmp != None:
	            strtmp = strtmp.strip("\n ")
	            infos["email"]=strtmp
	    span_address=soup.find('span', text="地址：")
	    if span_address != None:
	        strtmp=span_address.find_next_sibling().find_next().string
	        if strtmp != None:
	            strtmp = strtmp.strip("\n ")
	            infos["address"]=strtmp
	    # second block
	    td_time=soup.find('td', text=re.compile(r"成立日期"))
	    if td_time != None:
	        strtmp=td_time.find_next_sibling().string
	        if strtmp != None:
	            strtmp = strtmp.strip("\n ")
	            infos["time"]=strtmp
	    td_address=soup.find('td', text=re.compile(r"所属行业"))
	    if td_address != None:
	        strtmp=td_address.find_next_sibling().string
	        if strtmp != None:
	            strtmp = strtmp.strip("\n ")
	            infos["industry"]=strtmp
    except Exception as e:
        err=firm_name+":\n"+str(traceback.format_exc())
    finally:
        return infos, err

	
if __name__=="__main__":
    host="www.qichacha.com"
    headers = {'Accept-Encoding': 'gzip, deflate', 'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8', 'Connection': 'keep-alive', 'Host': 'www.qichacha.com', 'Accept-Language': 'zh-CN,zh;q=0.9', 'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.84 Safari/537.36'}
    cookies = get_cookie()
    # convert cookies dict to RequestsCookieJar object
    cookiejar=requests.utils.cookiejar_from_dict(cookies)

    s=requests.session()
    s.headers.update(headers)
    s.cookies=cookiejar

    # params 
    origin_sheet="data/sheet.xlsx"
    save_sheet="data/sheet-save.xlsx"
    
    wb = load_workbook(origin_sheet)
    sheet = wb.active
    
    # params
    begin = 3
    end = 10
    #count = 1622
    #count = 22
    c = 2 
	
    for r in range(begin,end+1):
        name = sheet.cell(row=r, column=c)
        #if r%50 == 0:
        print("row"+str(r)+": "+name.value)
	
	    # process
        infos, err  = get_infos(s, name.value)
        if err != "":
            print(err)
            break
	
	    # time
        cell = sheet.cell(row=r, column=5)
        cell.value = infos["time"]
	    # phone
        cell = sheet.cell(row=r, column=6)
        cell.value = infos["phone"]
	    # email
        cell = sheet.cell(row=r, column=7)
        cell.value = infos["email"]
	    # address
        cell = sheet.cell(row=r, column=8)
        cell.value = infos["address"]
	    # industry
        cell = sheet.cell(row=r, column=9)
        cell.value = infos["industry"]

        # one result, write once
        wb.save(save_sheet)
